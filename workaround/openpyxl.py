# -*- coding: utf-8 -*-
# SPDX-FileCopyrightText: 2025 Deutsche Telekom Technik GmbH <f.vonstudsinske@telekom.de>
# SPDX-License-Identifier: GPL-3.0-or-later

from pathlib import Path


def get_empty_workbook(keep_default_sheet: bool = True) -> 'openpyxl.workbook.Workbook':
    """ Workaround QGIS 3.34
        openpyxl: 3.1.2

        Workbook Erzeuging mit `openpyxl.Workbook()` bzw. `openpyxl.workbook.Workbook()`
        lässt QGIS abstürzen mit `Windows fatal exception: access violation`.
        In der Pythonkonsole kann normal ein leeres Workbook instanziert werden,
        jedoch teilweise nicht von der GUI aus.

        Workaround:
            - Leere vorher Excel-Datei abgespeichert
            - Excel-Datei wird mit `load_workbook` eingelesen
            - Vorhandene Sheets/Tabellenblätter werden gelöscht
            - Neues Tabellenblatt "Sheet" wird angelegt und als aktiv hinterlegt
            - Vom Dateisystem geladenes und überarbeitetes Workbook kann weiterverwendet werden

        :param keep_default_sheet: Keep a default and empty `Sheet` sheet. Defaults to True.

    """

    # import openpyxl
    from openpyxl import load_workbook

    # get the path to the empty Excel file
    empty_excel_file_path = Path(__file__).parent / "openpyxl_empty_workbook.xlsx"

    # load the empty workbook
    wb = load_workbook(empty_excel_file_path)

    # remove all sheets from the Excel file
    for sheet in list(wb.worksheets):
        wb.remove(sheet)

    # create a new sheet named "Sheet" and set it as active worksheet
    if keep_default_sheet:
        # create default/empty sheet to keep the default openpyxl Workbook behavior
        wb.create_sheet("Sheet")
        wb.active = wb["Sheet"]

    return wb


def df_to_rows(data_frame, index=True, header=True, skip_none_rows: bool = False):
    """ Workaround QGIS 3.34
        openpyxl: 3.1.2
        pandas: 2.2.2

        Pandas Dataframe kann mit "df.to_excel" nicht geschrieben werden.
        QGIS Crash `Windows fatal exception: access violation`.
        Workaround ist den Dataframe in openpyxl-Zeilen umzuwandenln.

        :param data_frame: Default argument for `openpyxl.utils.dataframe.dataframe_to_rows`
        :param index: Default argument for `openpyxl.utils.dataframe.dataframe_to_rows`
        :param header: Default argument for `openpyxl.utils.dataframe.dataframe_to_rows`
        :param skip_none_rows: Set to True to skip empty rows with only one value and the value is None
    """

    from openpyxl.utils.dataframe import dataframe_to_rows

    for row in dataframe_to_rows(data_frame, index, header):
        if skip_none_rows and len(row) == 1 and None in row:
            # skip empty row with only one value and the value is None
            continue
        else:
            yield row
